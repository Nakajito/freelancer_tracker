"""Proposal business-logic services: duplicate detection, status transitions,
and bulk import of scraper-extracted proposals."""

import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from pydantic import BaseModel

from apps.proposals.models import Client, Platform, Proposal, ProposalStatus, Tag


class DuplicateCheckResult(BaseModel):
    """Outcome of a duplicate-proposal check."""

    is_duplicate: bool
    existing_proposals: list = []


class StatusTransitionResult(BaseModel):
    """Outcome of a proposal status transition."""

    success: bool
    old_status: str
    new_status: str
    actual_response_date_set: bool = False


class DuplicateCheckService:
    """Detects whether a near-identical proposal was recently submitted."""

    @staticmethod
    def check_duplicate(
        owner,
        client,
        platform: str,
        days: int = 30,
    ) -> DuplicateCheckResult:
        """Check for proposals sent to the same client on the same platform recently.

        Args:
            owner: User who owns the proposals being checked.
            client: Client instance to match against.
            platform: Platform value (e.g. ``Platform.UPWORK``) to match.
            days: Look-back window in days. Defaults to 30.

        Returns:
            DuplicateCheckResult with ``is_duplicate=True`` and matching proposal
            ids/titles/dates when a duplicate exists; ``is_duplicate=False`` otherwise.
        """
        cutoff = timezone.now().date() - timedelta(days=days)

        existing = (
            Proposal.objects.for_user(owner)
            .with_client()
            .filter(
                client=client,
                platform=platform,
                sent_date__gte=cutoff,
            )
        )

        if existing.exists():
            return DuplicateCheckResult(
                is_duplicate=True,
                existing_proposals=list(existing.values("id", "title", "sent_date")),
            )

        return DuplicateCheckResult(is_duplicate=False)


class StatusTransitionService:
    """Applies validated status transitions to proposals and records response dates."""

    RESPONSE_STATUSES = {
        ProposalStatus.RESPONDED,
        ProposalStatus.NEGOTIATING,
        ProposalStatus.ACCEPTED,
        ProposalStatus.REJECTED,
    }

    @staticmethod
    def transition(
        proposal: Proposal, new_status: str, actor
    ) -> StatusTransitionResult:
        """Move a proposal to a new status, auto-setting ``actual_response_date``.

        If the target status is a response status (RESPONDED, NEGOTIATING, ACCEPTED,
        REJECTED) and ``actual_response_date`` has not yet been recorded, today's date
        is stamped automatically.

        Args:
            proposal: Proposal instance to update.
            new_status: One of the ``ProposalStatus`` string values.
            actor: User performing the transition (reserved for future audit log).

        Returns:
            StatusTransitionResult with the old/new status values and a flag
            indicating whether ``actual_response_date`` was set.
        """
        old_status = proposal.status
        response_date_set = False

        if (
            new_status in StatusTransitionService.RESPONSE_STATUSES
            and not proposal.actual_response_date
        ):
            proposal.actual_response_date = timezone.now().date()
            response_date_set = True

        proposal.status = new_status
        proposal.save()

        return StatusTransitionResult(
            success=True,
            old_status=old_status,
            new_status=new_status,
            actual_response_date_set=response_date_set,
        )


# --------------------------------------------------------------------------- #
# Bulk import (scraper file -> DRAFT proposals)
# --------------------------------------------------------------------------- #

# Map every accepted column/key (normalized to lowercase) to a model field.
# Covers both the JSON export keys (snake_case) and the CSV export labels.
_COLUMN_ALIASES = {
    "title": "title",
    "client": "client",
    "platform": "platform",
    "amount": "amount",
    "proposal_text": "proposal_text",
    "proposal text": "proposal_text",
    "status": "status",
    "sent_date": "sent_date",
    "sent date": "sent_date",
    "expected_response_date": "expected_response_date",
    "expected response": "expected_response_date",
    "expected response date": "expected_response_date",
    "actual_response_date": "actual_response_date",
    "actual response": "actual_response_date",
    "actual response date": "actual_response_date",
    "paid": "paid",
    "tags": "tags",
    "job_url": "job_url",
    "job url": "job_url",
    "proposal_url": "proposal_url",
    "proposal url": "proposal_url",
}

_SUPPORTED_EXTENSIONS = {".json", ".csv", ".xlsx", ".md"}
_TRUTHY = {"yes", "true", "1", "y", "si", "sí"}


@dataclass
class ImportResult:
    """Outcome of a bulk import."""

    created: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def _normalize_key(key) -> str | None:
    """Map a raw column/key to a model field name, or ``None`` if unknown."""
    if key is None:
        return None
    return _COLUMN_ALIASES.get(str(key).strip().lower())


def _normalize_row(raw: dict) -> dict:
    """Keep only recognized keys, mapped to model field names."""
    row: dict = {}
    for key, value in raw.items():
        field_name = _normalize_key(key)
        if field_name is not None:
            row[field_name] = value
    return row


class ProposalImportService:
    """Parse an uploaded file and create DRAFT proposals from its rows.

    Accepts the same field shape the export produces (round-trip) across
    ``.json``, ``.csv``, ``.xlsx`` and a Markdown table (``.md``). Status from
    the file is always overridden to ``DRAFT`` so everything lands for review.
    """

    SUPPORTED_EXTENSIONS = _SUPPORTED_EXTENSIONS

    # Every row costs a full_clean() plus an INSERT in the request thread, and
    # a 5 MB .xlsx can expand to millions of rows. Cap the batch rather than
    # let one upload hold a gunicorn worker (there are only three) indefinitely.
    MAX_ROWS = 5_000

    # ----------------------------- parsing ----------------------------- #
    @classmethod
    def parse(cls, uploaded_file) -> list[dict]:
        """Return a list of normalized row dicts from ``uploaded_file``.

        Raises ``ValueError`` for unsupported extensions or malformed files.
        """
        suffix = Path(uploaded_file.name).suffix.lower()
        raw_bytes = uploaded_file.read()

        if suffix == ".json":
            rows = cls._parse_json(raw_bytes)
        elif suffix == ".csv":
            rows = cls._parse_csv(raw_bytes)
        elif suffix == ".xlsx":
            rows = cls._parse_xlsx(raw_bytes)
        elif suffix == ".md":
            rows = cls._parse_md(raw_bytes)
        else:
            raise ValueError(
                f"Unsupported file type '{suffix}'. "
                f"Use one of: {', '.join(sorted(_SUPPORTED_EXTENSIONS))}."
            )

        if len(rows) > cls.MAX_ROWS:
            raise ValueError(
                f"File contains {len(rows)} rows; the maximum is {cls.MAX_ROWS}. "
                "Split it into smaller files."
            )

        return [_normalize_row(r) for r in rows]

    @staticmethod
    def _parse_json(raw: bytes) -> list[dict]:
        try:
            data = json.loads(raw.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise ValueError(f"Invalid JSON file: {exc}") from exc
        if isinstance(data, dict):
            # accept either a single proposal or a wrapper {"proposals": [...]}
            if isinstance(data.get("proposals"), list):
                data = data["proposals"]
            else:
                data = [data]
        if not isinstance(data, list):
            raise ValueError("JSON must be a list of proposals or a single object.")
        return [d for d in data if isinstance(d, dict)]

    @staticmethod
    def _parse_csv(raw: bytes) -> list[dict]:
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]

    @staticmethod
    def _parse_xlsx(raw: bytes) -> list[dict]:
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # openpyxl raises a variety of errors
            raise ValueError(f"Invalid XLSX file: {exc}") from exc
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header]
        result: list[dict] = []
        for values in rows_iter:
            if len(result) > ProposalImportService.MAX_ROWS:
                # Stop reading rather than materialise an unbounded list; parse()
                # rejects the file once it sees the count exceeded.
                break
            if values is None or all(v is None for v in values):
                continue
            row = {}
            for key, value in zip(headers, values):
                if not key:
                    continue
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()[:10]
                row[key] = value
            result.append(row)
        return result

    @staticmethod
    def _parse_md(raw: bytes) -> list[dict]:
        text = raw.decode("utf-8")
        pipe_lines = [
            line.strip() for line in text.splitlines() if line.strip().startswith("|")
        ]
        if len(pipe_lines) < 2:
            return []

        def cells(line: str) -> list[str]:
            return [c.strip() for c in line.strip().strip("|").split("|")]

        headers = cells(pipe_lines[0])
        separator = re.compile(r"^[\s:\-]+$")
        data_lines = [
            line
            for line in pipe_lines[1:]
            if not all(separator.match(c) for c in cells(line))
        ]
        result: list[dict] = []
        for line in data_lines:
            values = cells(line)
            result.append(dict(zip(headers, values)))
        return result

    # ----------------------------- importing --------------------------- #
    @classmethod
    def import_file(cls, user, uploaded_file) -> ImportResult:
        """Parse ``uploaded_file`` and create DRAFT proposals for ``user``."""
        rows = cls.parse(uploaded_file)
        return cls.import_rows(user, rows)

    @classmethod
    def import_rows(cls, user, rows: list[dict]) -> ImportResult:
        result = ImportResult()
        for index, row in enumerate(rows, start=1):
            title = str(row.get("title") or "").strip()
            label = title or f"row {index}"
            try:
                created = cls._import_one(user, row, title)
            except (ValidationError, ValueError) as exc:
                result.errors.append(f"{label}: {cls._error_message(exc)}")
                continue
            if created:
                result.created += 1
            else:
                result.skipped += 1
        return result

    @classmethod
    @transaction.atomic
    def _import_one(cls, user, row: dict, title: str) -> bool:
        """Create one proposal. ``True`` if created, ``False`` if skipped.

        Wrapped in a savepoint so a failing row never aborts the batch.
        """
        if not title:
            raise ValueError("missing required title")

        client_name = str(row.get("client") or "").strip()

        if cls._is_duplicate(user, title, client_name):
            return False

        client = None
        if client_name:
            client, _ = Client.objects.get_or_create(owner=user, name=client_name)

        proposal = Proposal(
            owner=user,
            title=title,
            client=client,
            status=ProposalStatus.DRAFT,
            platform=cls._coerce_platform(row.get("platform")),
            amount=cls._coerce_decimal(row.get("amount")),
            proposal_text=str(row.get("proposal_text") or ""),
            sent_date=cls._coerce_date(row.get("sent_date")),
            expected_response_date=cls._coerce_date(row.get("expected_response_date")),
            actual_response_date=cls._coerce_date(row.get("actual_response_date")),
            paid=cls._coerce_bool(row.get("paid")),
            job_url=str(row.get("job_url") or ""),
            proposal_url=str(row.get("proposal_url") or ""),
        )
        proposal.full_clean(exclude=["client"])
        proposal.save()

        for tag_name in cls._coerce_tags(row.get("tags")):
            tag, _ = Tag.objects.get_or_create(
                owner=user,
                slug=slugify(tag_name)[:50],
                defaults={"name": tag_name},
            )
            proposal.tags.add(tag)

        return True

    @staticmethod
    def _is_duplicate(user, title: str, client_name: str) -> bool:
        qs = Proposal.objects.for_user(user).filter(title__iexact=title)
        if client_name:
            qs = qs.filter(client__name__iexact=client_name)
        else:
            qs = qs.filter(client__isnull=True)
        return qs.exists()

    # ----------------------------- coercion ---------------------------- #
    @staticmethod
    def _coerce_platform(raw) -> str:
        if not raw:
            return Platform.OTHER
        value = str(raw).strip().lower()
        for choice_value, label in Platform.choices:
            if value in (choice_value.lower(), str(label).lower()):
                return choice_value
        return Platform.OTHER

    @staticmethod
    def _coerce_decimal(raw) -> Decimal:
        if raw is None or raw == "":
            return Decimal("0.00")
        if isinstance(raw, bool):  # guard: bool is an int subclass
            raise ValueError("invalid amount")
        if isinstance(raw, (int, float, Decimal)):
            return Decimal(str(raw))
        cleaned = str(raw).strip().replace("$", "").replace(",", "")
        if not cleaned:
            return Decimal("0.00")
        try:
            return Decimal(cleaned)
        except InvalidOperation as exc:
            raise ValueError(f"invalid amount '{raw}'") from exc

    @staticmethod
    def _coerce_date(raw):
        if raw is None or raw == "":
            return None
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        text = str(raw).strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError as exc:
            raise ValueError(f"invalid date '{raw}' (use YYYY-MM-DD)") from exc

    @staticmethod
    def _coerce_bool(raw) -> bool:
        if isinstance(raw, bool):
            return raw
        if raw is None:
            return False
        return str(raw).strip().lower() in _TRUTHY

    @staticmethod
    def _coerce_tags(raw) -> list[str]:
        if not raw:
            return []
        if isinstance(raw, (list, tuple)):
            items = [str(t) for t in raw]
        else:
            items = re.split(r"[;,]", str(raw))
        return [t.strip() for t in items if t.strip()]

    @staticmethod
    def _error_message(exc: Exception) -> str:
        if isinstance(exc, ValidationError):
            return "; ".join(
                f"{f}: {' '.join(map(str, msgs))}"
                for f, msgs in exc.message_dict.items()
            )
        return str(exc)
